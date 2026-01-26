import asyncio
import os
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from backend.settings import get_settings


settings = get_settings()


def _dedupe_columns(columns: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for name in columns:
        base = name.strip() if name else "Column"
        count = seen.get(base, 0)
        if count:
            new_name = f"{base}_{count + 1}"
        else:
            new_name = base
        seen[base] = count + 1
        result.append(new_name)
    return result


def _column_key_from_letter(letter: str) -> int | None:
    try:
        return column_index_from_string(letter.strip().upper())
    except Exception:
        return None


def _coerce_date(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%y", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt)
            except ValueError:
                continue
    return None


def _coerce_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        raw = value.strip().replace(",", "").replace("$", "")
        try:
            return float(raw)
        except ValueError:
            return None
    return None


def resolve_data_path(path: str) -> Path:
    raw = path.strip()
    if raw.startswith("\\\\") or raw.startswith("//"):
        return Path(raw)
    if os.path.isabs(raw):
        return Path(raw)
    return Path(settings.data_root).joinpath(raw)


@dataclass
class ExcelCacheEntry:
    columns: list[str]
    rows: list[dict[str, Any]]
    mtime: float
    cached_at: float


def _read_excel_rows(path: Path, sheet_name: str | None, has_header: bool) -> ExcelCacheEntry:
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        if sheet_name:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        if has_header:
            header_row = next(rows_iter, None)
            if not header_row:
                columns: list[str] = []
                data_rows: list[dict[str, Any]] = []
            else:
                columns = _dedupe_columns(
                    [str(cell).strip() if cell is not None else "" for cell in header_row]
                )
                data_rows = []
                for row in rows_iter:
                    if not row or all(cell is None for cell in row):
                        continue
                    row_dict = {
                        columns[idx]: row[idx] if idx < len(row) else None
                        for idx in range(len(columns))
                    }
                    data_rows.append(row_dict)
        else:
            max_cols = worksheet.max_column or 0
            columns = [get_column_letter(idx) for idx in range(1, max_cols + 1)]
            data_rows = []
            for row in rows_iter:
                if not row or all(cell is None for cell in row):
                    continue
                row_dict = {
                    columns[idx]: row[idx] if idx < len(row) else None
                    for idx in range(len(columns))
                }
                data_rows.append(row_dict)
    finally:
        workbook.close()
    mtime = os.path.getmtime(path)
    return ExcelCacheEntry(columns=columns, rows=data_rows, mtime=mtime, cached_at=time.time())


class ExcelCache:
    def __init__(self, ttl_seconds: int = 300) -> None:
        self._ttl_seconds = ttl_seconds
        self._cache: dict[str, ExcelCacheEntry] = {}
        self._lock = asyncio.Lock()

    async def get_rows(
        self,
        path: Path,
        sheet_name: str | None,
        has_header: bool,
        force: bool = False,
    ) -> ExcelCacheEntry:
        key = f"{path}::{sheet_name or ''}::{has_header}"
        if not force:
            async with self._lock:
                entry = self._cache.get(key)
                if entry:
                    try:
                        mtime = os.path.getmtime(path)
                    except FileNotFoundError:
                        self._cache.pop(key, None)
                    else:
                        if mtime == entry.mtime and (time.time() - entry.cached_at) < self._ttl_seconds:
                            return entry
        entry = await asyncio.to_thread(_read_excel_rows, path, sheet_name, has_header)
        async with self._lock:
            self._cache[key] = entry
        return entry


def resolve_mapping(columns: list[str], excel_column: str) -> str | None:
    raw = excel_column.strip()
    if not raw:
        return None
    letter_index = _column_key_from_letter(raw)
    if letter_index is not None and 1 <= letter_index <= len(columns):
        return columns[letter_index - 1]
    for col in columns:
        if col.strip().lower() == raw.lower():
            return col
    return None


def map_rows(
    rows: list[dict[str, Any]],
    columns: list[str],
    mappings: list[dict[str, str]],
) -> list[dict[str, Any]]:
    resolved: list[tuple[str, str, str]] = []
    for mapping in mappings:
        column_key = resolve_mapping(columns, mapping["excel_column"])
        if not column_key:
            continue
        resolved.append((mapping["field_name"], column_key, mapping.get("data_type", "text")))

    normalized: list[dict[str, Any]] = []
    for row in rows:
        item: dict[str, Any] = {}
        for field_name, column_key, data_type in resolved:
            value = row.get(column_key)
            if data_type == "date":
                value = _coerce_date(value)
            elif data_type == "number":
                value = _coerce_number(value)
            item[field_name] = value
        normalized.append(item)
    return normalized


excel_cache = ExcelCache(ttl_seconds=settings.excel_cache_ttl_seconds)
